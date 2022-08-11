from WebScrapers.scraper1 import site1Products
from WebScrapers.scraper2 import site2Products
from WebScrapers.scraper3 import site3Products
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def main():
    print("For faster results please be specific with your search.")

    product = input("\033[96m {}\033[00m" .format(
        "What product are you searching for? "))
    print("Running script...")

    sorted_products = getProducts(product)

    if len(sorted_products) > 0:
        export(sorted_products)
        print("\033[92m {}\033[00m".format(
            f"{len(sorted_products)}" + " exported to Excel"))


def getProducts(product):
    total_found_products = {}

    products1 = site1Products(product)
    products2 = site2Products(product)
    products3 = site3Products(product)

    if(products1):
        for product in products1:
            total_found_products[product] = products1[product]
    else:
        print("Product not found on Newegg")

    if(products2):
        for product in products2:
            total_found_products[product] = products2[product]
    else:
        print("Product not found on Microcenter")

    if(products3 and len(products3) != 0):
        for product in products3:
            total_found_products[product] = products3[product]
    else:
        print("Product not found on Memoryexpress")

    sorted_products = sorted(total_found_products.items(),
                             key=lambda x: x[1]['price'])

    return sorted_products


def export(sorted_products):

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    ws.merge_cells('A1:D2')
    ws.merge_cells('A3:C3')
    ws['A1'] = "Lowest to Highest Price"
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(
        start_color='A5C249', end_color='A5C249', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = "Name/Link"
    ws['A3'].fill = PatternFill(
        start_color='C8DA92', end_color='C8DA92', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D3'] = "Price"
    ws['D3'].fill = PatternFill(
        start_color='C8DA92', end_color='C8DA92', fill_type='solid')
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')

    row = 4
    for product in sorted_products:
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = product[0]
        ws[f"A{row}"].hyperlink = product[1]["link"]
        ws[f"A{row}"].alignment = Alignment(
            horizontal='left', vertical='center')
        ws[f"A{row}"].font = Font(underline='single')
        ws[f"D{row}"] = product[1]["price"]
        ws[f"D{row}"].alignment = Alignment(
            horizontal='right', vertical='center')
        row += 1

    ws.merge_cells("F1:G2")
    ws.merge_cells("H1:I2")
    ws.merge_cells("F3:G3")
    ws.merge_cells("H3:I3")
    ws['F1'] = "# of Products"
    ws['F1'].font = Font(bold=True)
    ws['F1'].fill = PatternFill(
        start_color='A5C249', end_color='A5C249', fill_type='solid')
    ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'] = "Average Price"
    ws['H1'].font = Font(bold=True)
    ws['H1'].fill = PatternFill(
        start_color='A5C249', end_color='A5C249', fill_type='solid')
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F3'] = int(len(sorted_products))
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H3'].alignment = Alignment(horizontal='center', vertical='center')

    total = 0
    for product in sorted_products:
        total += int(product[1]["price"])

    total = total / len(sorted_products)
    ws['H3'] = int(total)

    wb.save("data.xlsx")


main()
